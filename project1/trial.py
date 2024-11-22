from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Add some data to the sheet
data = [["ID", "Name"], [1, "Alice"], [2, "Bob"], [3, "Charlie"], [4, "David"]]

for row in data:
    ws.append(row)

# Define the yellow fill
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Apply the yellow fill to the cells in the "ID" column where ID is 2 or 4
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    if row[0].value in [2, 4]:
        row[0].fill = yellow_fill

# Save the workbook
wb.save("output/toy_example.xlsx")
