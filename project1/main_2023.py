from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from project1.utils import read_excel

results_df = read_excel(path="data/Startups Flanders 2023.xlsx", sheet_name="Results")
sheet1_df = read_excel(path="data/Startups Flanders 2023.xlsx", sheet_name="Sheet1")

column1 = "ID"
column2 = "ID"

# Find indices of ID numbers in Results that are in Sheet 1
common_indices = results_df[results_df[column1].isin(sheet1_df[column2])].index

# Find indices of IDs that are in Sheet 1 but not in Results
missing_indices = sheet1_df[~sheet1_df[column2].isin(results_df[column1])].index

# Print the number of common IDs
print(f"Number of common IDs: {len(common_indices)}")
print(f"Number of IDs in Sheet1 but not in Results: {len(missing_indices)}")

# Save the DataFrame to Excel
output_path = "output/Startups Flanders 2023 Arda.xlsx"
results_df.to_excel(output_path, sheet_name="Results", index=False)

# Save the missing IDs to a different Excel file
missing_ids_output_path = "output/To Be Added To 2023.xlsx"
sheet1_df.loc[missing_indices].to_excel(
    missing_ids_output_path, sheet_name="Missing IDs", index=False
)

# Load the workbook and select the sheet
wb = load_workbook(output_path)
ws = wb["Results"]

# Define the yellow and red fills
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Apply the yellow fill to the cells in the "ID" column with common indices
for idx in common_indices:
    cell = ws.cell(
        row=idx + 2, column=1
    )  # +2 because DataFrame index is 0-based and Excel is 1-based, plus header row
    cell.fill = yellow_fill

# Save the workbook
wb.save(output_path)

# Highlight missing IDs in the original sheet1_df
sheet1_output_path = "output/Sheet1_Highlighted.xlsx"
sheet1_df.to_excel(sheet1_output_path, sheet_name="Sheet1", index=False)

# Load the workbook and select the sheet
wb_sheet1 = load_workbook(sheet1_output_path)
ws_sheet1 = wb_sheet1["Sheet1"]

# Apply the red fill to the cells in the "ID" column with missing indices
for idx in missing_indices:
    cell = ws_sheet1.cell(
        row=idx + 2, column=1
    )  # +2 because DataFrame index is 0-based and Excel is 1-based, plus header row
    cell.fill = red_fill

# Save the workbook
wb_sheet1.save(sheet1_output_path)
